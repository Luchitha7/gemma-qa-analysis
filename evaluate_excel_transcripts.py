"""
Batch QA Evaluation Script for Excel Transcripts.

Reads transcripts from 'qa_evaluation_transcripts_100.xlsx', sends POST requests to
http://localhost:8000/api/tenants/{tenant_id}/evaluate one by one, and updates the Excel
sheet with the received score in a new column.

Color Highlights:
- Green: Very close to expected score (|actual - expected| <= 5.0)
- Yellow: Bit far (|actual - expected| <= 15.0)
- Red: Far (|actual - expected| > 15.0 or request error)
"""

import sys
import os
import time
import json
import argparse
import urllib.request
import urllib.error
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Color palette for highlights
GREEN_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")  # Light Green
GREEN_FONT = Font(name="Calibri", size=11, bold=True, color="166534")

YELLOW_FILL = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid") # Light Yellow
YELLOW_FONT = Font(name="Calibri", size=11, bold=True, color="854D0E")

RED_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")    # Light Red
RED_FONT = Font(name="Calibri", size=11, bold=True, color="991B1B")

HEADER_FILL = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark Blue
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

THIN_BORDER = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)


def send_evaluate_request(endpoint: str, transcript: str, channel: str = "Call", agent_name: str = "Agent", timeout: int = 300) -> dict:
    """Send evaluation POST request to FastAPI server using standard library urllib."""
    payload = {
        "transcript": transcript,
        "channel": channel or "Call",
        "agent_name": agent_name or "Agent"
    }
    data_bytes = json.dumps(payload).encode("utf-8")
    
    req = urllib.request.Request(
        endpoint,
        data=data_bytes,
        headers={"Content-Type": "application/json", "Accept": "application/json"},
        method="POST"
    )
    
    with urllib.request.urlopen(req, timeout=timeout) as response:
        resp_text = response.read().decode("utf-8")
        return json.loads(resp_text)


def evaluate_excel_rows(
    excel_path: str = "qa_evaluation_transcripts_100.xlsx",
    endpoint: str = "http://localhost:8000/api/tenants/S-NET/evaluate",
    start_row: int = 2,
    max_rows: int = None,
    delay: float = 0.5,
    timeout: int = 300
):
    """Iterate through Excel rows, send evaluation requests, track response time, and update the workbook."""
    if not os.path.exists(excel_path):
        print(f"Error: Excel file '{excel_path}' not found.")
        return

    print("=" * 80)
    print("🚀 Starting Batch QA Server Evaluation with Latency Tracking")
    print(f"📁 Target Excel File:  {excel_path}")
    print(f"🌐 Server Endpoint:   {endpoint}")
    print(f"⏱️ Request Timeout:   {timeout} seconds (handles up to {timeout // 60}m responses)")
    print("=" * 80)

    wb = openpyxl.load_workbook(excel_path)
    sheet_name = "QA Interaction Transcripts" if "QA Interaction Transcripts" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    # Target column for Actual Server Score is column 10 (J)
    score_col = 10
    score_col_letter = get_column_letter(score_col)

    header_score_cell = ws.cell(row=1, column=score_col)
    header_score_cell.value = "Server QA Score (Actual)"
    header_score_cell.fill = HEADER_FILL
    header_score_cell.font = HEADER_FONT
    header_score_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_score_cell.border = THIN_BORDER
    ws.column_dimensions[score_col_letter].width = 18

    # Target column for Response Time is column 11 (K)
    time_col = 11
    time_col_letter = get_column_letter(time_col)

    header_time_cell = ws.cell(row=1, column=time_col)
    header_time_cell.value = "Response Time (s)"
    header_time_cell.fill = HEADER_FILL
    header_time_cell.font = HEADER_FONT
    header_time_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_time_cell.border = THIN_BORDER
    ws.column_dimensions[time_col_letter].width = 16

    total_rows = ws.max_row
    end_row = total_rows if max_rows is None else min(total_rows, start_row + max_rows - 1)
    
    success_count = 0
    error_count = 0
    green_count = 0
    yellow_count = 0
    red_count = 0
    latencies = []

    for r in range(start_row, end_row + 1):
        record_id = ws.cell(row=r, column=1).value or f"Row-{r}"
        scenario_title = ws.cell(row=r, column=2).value or ""
        channel = ws.cell(row=r, column=3).value or "Call"
        expected_score = ws.cell(row=r, column=6).value
        transcript = ws.cell(row=r, column=9).value

        if not transcript:
            print(f"⚠️ [Row {r}/{total_rows}] {record_id}: Empty transcript, skipping.")
            continue

        try:
            expected_score_float = float(expected_score) if expected_score is not None else 0.0
        except (ValueError, TypeError):
            expected_score_float = 0.0

        print(f"\n⏳ [{r - start_row + 1}/{end_row - start_row + 1}] Sending {record_id}: '{scenario_title[:32]}...' (Expected: {expected_score_float:.1f})")

        score_cell = ws.cell(row=r, column=score_col)
        score_cell.border = THIN_BORDER
        score_cell.alignment = Alignment(horizontal="center", vertical="top")

        time_cell = ws.cell(row=r, column=time_col)
        time_cell.border = THIN_BORDER
        time_cell.alignment = Alignment(horizontal="center", vertical="top")
        time_cell.font = Font(name="Calibri", size=10)

        start_time = time.perf_counter()

        try:
            res = send_evaluate_request(endpoint, transcript=transcript, channel=channel, timeout=timeout)
            elapsed_sec = round(time.perf_counter() - start_time, 2)
            latencies.append(elapsed_sec)

            actual_score = float(res.get("final_score", 0.0))
            diff = abs(actual_score - expected_score_float)

            # Insert score into Column J
            score_cell.value = actual_score
            score_cell.number_format = '0.0'

            # Insert response time into Column K
            time_cell.value = elapsed_sec
            time_cell.number_format = '0.0 "s"'

            # Apply Color Highlighting Rule:
            # - Green: Very close (diff <= 5.0)
            # - Yellow: Bit far (5.0 < diff <= 15.0)
            # - Red: Far (diff > 15.0)
            if diff <= 5.0:
                score_cell.fill = GREEN_FILL
                score_cell.font = GREEN_FONT
                badge = "🟢 GREEN (Very Close)"
                green_count += 1
            elif diff <= 15.0:
                score_cell.fill = YELLOW_FILL
                score_cell.font = YELLOW_FONT
                badge = "🟡 YELLOW (Bit Far)"
                yellow_count += 1
            else:
                score_cell.fill = RED_FILL
                score_cell.font = RED_FONT
                badge = "🔴 RED (Far Difference)"
                red_count += 1

            mins = int(elapsed_sec // 60)
            secs = elapsed_sec % 60
            time_str = f"{mins}m {secs:.1f}s" if mins > 0 else f"{secs:.2f}s"

            print(f"   ✅ Server Score: {actual_score:.1f} | Expected: {expected_score_float:.1f} | Diff: {diff:.1f} -> {badge}")
            print(f"   ⏱️ Response Time: {time_str} ({elapsed_sec:.2f}s)")
            success_count += 1

        except urllib.error.URLError as ue:
            elapsed_sec = round(time.perf_counter() - start_time, 2)
            print(f"   ❌ Network/HTTP Error on {record_id} after {elapsed_sec:.2f}s: {ue.reason}")
            score_cell.value = "ERR"
            score_cell.fill = RED_FILL
            score_cell.font = RED_FONT
            time_cell.value = elapsed_sec
            time_cell.number_format = '0.0 "s"'
            error_count += 1
            red_count += 1
        except Exception as e:
            elapsed_sec = round(time.perf_counter() - start_time, 2)
            print(f"   ❌ Evaluation Error on {record_id} after {elapsed_sec:.2f}s: {str(e)}")
            score_cell.value = "ERR"
            score_cell.fill = RED_FILL
            score_cell.font = RED_FONT
            time_cell.value = elapsed_sec
            time_cell.number_format = '0.0 "s"'
            error_count += 1
            red_count += 1

        # Auto-save workbook after every row to protect progress
        wb.save(excel_path)

        if delay > 0:
            time.sleep(delay)

    avg_latency = (sum(latencies) / len(latencies)) if latencies else 0.0
    min_latency = min(latencies) if latencies else 0.0
    max_latency = max(latencies) if latencies else 0.0

    print("\n" + "=" * 80)
    print("📊 Evaluation Run Summary:")
    print(f"   • Total Processed: {end_row - start_row + 1}")
    print(f"   • Successful:      {success_count}")
    print(f"   • Errors:          {error_count}")
    print(f"   • 🟢 Green (Very Close, <= 5 pts):  {green_count}")
    print(f"   • 🟡 Yellow (Bit Far, 5-15 pts):    {yellow_count}")
    print(f"   • 🔴 Red (Far / Error, > 15 pts):   {red_count}")
    print("⏱️ Latency Statistics:")
    print(f"   • Average Response Time: {avg_latency:.2f}s")
    print(f"   • Min Response Time:     {min_latency:.2f}s")
    print(f"   • Max Response Time:     {max_latency:.2f}s")
    print(f"💾 Updated Excel Saved to: {excel_path}")
    print("=" * 80)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Evaluate QA interaction transcripts from Excel against QA Score Server.")
    parser.add_argument("--excel", default="qa_evaluation_transcripts_100.xlsx", help="Path to input Excel file.")
    parser.add_argument("--endpoint", default="http://localhost:8000/api/tenants/S-NET/evaluate", help="QA server evaluation endpoint URL.")
    parser.add_argument("--start-row", type=int, default=2, help="Row number to start evaluation from (default: 2).")
    parser.add_argument("--max-rows", type=int, default=None, help="Maximum number of rows to evaluate (default: all).")
    parser.add_argument("--delay", type=float, default=0.3, help="Delay in seconds between requests (default: 0.3s).")
    parser.add_argument("--timeout", type=int, default=300, help="HTTP request timeout in seconds (default: 300s / 5 min).")

    args = parser.parse_args()

    evaluate_excel_rows(
        excel_path=args.excel,
        endpoint=args.endpoint,
        start_row=args.start_row,
        max_rows=args.max_rows,
        delay=args.delay,
        timeout=args.timeout
    )
